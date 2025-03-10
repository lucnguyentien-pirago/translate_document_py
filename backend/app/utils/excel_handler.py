import openpyxl
import io
from typing import List, Dict, Tuple

class ExcelHandler:
    def __init__(self):
        pass
    
    def extract_text_from_excel(self, file_content: bytes) -> List[Dict[str, List[Dict[str, str]]]]:
        """
        Trích xuất văn bản từ file Excel
        Trả về danh sách các sheet, mỗi sheet chứa danh sách các cell
        """
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            result = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                cells = []
                
                # Duyệt qua các ô không trống trong sheet
                for row in range(1, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=row, col=col).value
                        if cell_value:
                            cell_address = f"{openpyxl.utils.get_column_letter(col)}{row}"
                            cells.append({
                                "address": cell_address,
                                "content": str(cell_value)
                            })
                
                result.append({
                    "sheet_name": sheet_name,
                    "cells": cells
                })
                
            return result
        except Exception as e:
            raise Exception(f"Lỗi khi xử lý file Excel: {str(e)}")
    
    def create_translated_excel(self, original_file: bytes, translated_data: List[Dict[str, List[Dict[str, str]]]]) -> bytes:
        """
        Tạo file Excel với nội dung đã được dịch
        """
        try:
            # Đọc file Excel gốc
            workbook = openpyxl.load_workbook(io.BytesIO(original_file))
            
            # Cập nhật nội dung đã dịch vào workbook
            for sheet_data in translated_data:
                sheet_name = sheet_data["sheet_name"]
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    for cell_data in sheet_data["cells"]:
                        address = cell_data["address"]
                        translated_content = cell_data.get("translated_content", "")
                        if translated_content:
                            sheet[address] = translated_content
            
            # Lưu workbook vào buffer
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output.getvalue()
        except Exception as e:
            raise Exception(f"Lỗi khi tạo file Excel đã dịch: {str(e)}") 